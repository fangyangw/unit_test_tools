#coding: utf-8
import sys
import uuid
import random
import datetime
import json
import os
import subprocess
import re
import shutil
import time
import chardet
import threading
import xlrd
import openpyxl
from xml.dom.minidom import parse
from mypackages.zlims_tools import ZLIMSClient, MqClient, logging
from config import TEST_PRODUCTS, CHECK_INSTRUMENT_CURRENT_STATUS, CHECK_RSYNC_STATUS, TEST_LANGUAGE, TEST_CAL_FILE, WRITE_FASTQ_WDL_PATH, basecall_uid, basecall_gid
from sys import version_info

logger = logging.getLogger("Tools")

if version_info.major == 2:
    reload(sys)
    sys.setdefaultencoding('utf-8')

BARCODE_START = 0
ONE_SAMPLE_BARCODE_NUMBER = 1
BARCODE_LIST = [i for i in range(1, 129)]
BARCODE_LIST.extend([i for i in range(501, 597)])
double_barcodes = []
for b in BARCODE_LIST:
    for i in BARCODE_LIST:
        double_barcodes.append("%s_%s" % (b, i))
BARCODE_LIST.extend(double_barcodes)

ANALYSIS_TASK_START_WAIT_TIME = 3 * 60
CYCLE_LENGTH = 5
TEST_FASTQ_PATH = "test_fastq"
T7 = "DNBSEQ-T7"
ONLY_SUPPORT_ONE_LANE = ["FIS"]
CONFIG_MAP = {
    "MGISEQ-50": {
        "config_file": "C:\\BGI\\Config\\BGI.ZebraV01Seq.Service.xml"
    },
    "MGISEQ-200": {
        "config_file": "C:\\BGI\\Config\\BGI.ZebraV01Seq.Service.xml"
    },
    "DNBSEQ-G50": {
        "config_file": "C:\\BGI\\Config\\BGI.ZebraV01Seq.Service.xml"
    },
    "MGISEQ-2000": {
        "config_file": "C:\\BGI\\Config\\BGI.ZebraV2Seq.Service.xml"
    },
    "DNBSEQ-G400": {
        "config_file": "C:\\BGI\\Config\\BGI.ZebraV2Seq.Service.xml"
    },
    "DNBSEQ-T7": {
        "config_file": "C:\\BGI\\Config\\BGI.ZebraV40Seq.Service.xml"
    }
}

def get_machine_sn():
    return os.environ.get('COMPUTERNAME', 'SZMGIPB591D')

def getValueByTagName(ele, tag_name):
    elements = ele.getElementsByTagName(tag_name)
    if elements:
        if elements[0].childNodes:
            return elements[0].childNodes[0].data
        else:
            return None
    else:
        return None

def getParaNameDict(items):
    para_name_dict = {}
    for item in items:
        para_set_name = getValueByTagName(item, 'Para_Set_Name')
        if not para_name_dict.get(para_set_name):
            para_name_dict[para_set_name] = {}
        para_name = getValueByTagName(item, 'Para_Name')
        default_value = getValueByTagName(item, 'Def_Value')
        currrent_value = getValueByTagName(item, 'Cur_Value')
        instrument_name = getValueByTagName(item, 'Instrument_Name')
        if currrent_value:
            value = currrent_value
        else:
            value = default_value
        if para_name_dict.get(para_name):
            print(para_name + ' has dup value!')
        para_name_dict[para_set_name][para_name] = value
        para_name_dict[para_set_name]['Instrument_Name'] = instrument_name
    return para_name_dict


def get_analysis_time(product):
    maps = {
        "QC": 5 * 60,
        "BFI": 15 * 60,
        "PFI": 5 * 60,
        "WGS": 20 * 60,
        "WES": 12 * 60,
        "AccuR-Seq_PGS": 20 * 60,
        "AccuR-Seq_NIPT": 5 * 60,
        "write_fastq": 10 * 60,
    }
    analysis_time = maps.get(product, 15 * 60)
    return analysis_time

def notification_console(message):
    try:
        cmd = "echo %s" % message
        if os.name == 'nt':
            cmd = "@" + cmd
        os.system(cmd)
    except:
        pass

class ZLIMSServiceError(Exception):
    pass


class WholePipelineTest:

    def __init__(self, product, sequencing_mode=None, language=TEST_LANGUAGE):
        self.product = product
        self.sequencing_mode = sequencing_mode
        self.zlims_client = ZLIMSClient()
        self.machine_sn = get_machine_sn()
        self.history_choice_dict = {}
        self.samples = []
        self.warnings = []
        self.need_check_values = {}
        self.mq_client = None
        self.mq_listen_thread = None
        self.language = language
        self.test_time = datetime.datetime.utcnow()

    def parse_software_conf(self):
        config = CONFIG_MAP.get(self.machine_part_number, {})
        config_file = config.get("config_file")
        if not os.path.exists(config_file):
            raise Exception("Error: 找不到控制软件配置文件，请检查控制软件是否安装正常。")

        dom_tree = parse(config_file)
        collection = dom_tree.documentElement
        ii = collection.getElementsByTagName('InstrumentConfig')
        para_name_dict = getParaNameDict(ii)

        return para_name_dict

    def import_dnb_sample_to_lims(self, number, one_sample_barcode_number=ONE_SAMPLE_BARCODE_NUMBER):
        if self.language == 'cn':
            template_excel = os.path.join(sys.path[0], "zlims_sample_excel", self.language, "%s2020-05-18.xlsx" %self.product)
        else:
            template_excel = os.path.join(sys.path[0], "zlims_sample_excel", self.language, "Sample_import_template%s2020-05-18.xlsx" % self.product)
        import_excel = "import_temp.xlsx"
        if os.path.exists(import_excel):
            os.remove(import_excel)
        data = xlrd.open_workbook(template_excel)
        table = data.sheets()[0]
        try:
            if len(table.row_values(0)) < 2:
                table = data.sheets()[1]
        except:
            table = data.sheets()[1]
        sample_id_col, dnb_id_col, barcode_col = 0, 0, 0

        wb = openpyxl.Workbook()
        sheet = wb.active
        if self.language == 'cn':
            sheet.title = 'DNB样本录入'
        else:
            sheet.title = 'DNB Sample Entry'
        for i, header in enumerate(table.row_values(0)):
            sheet.cell(row=1, column=i+1, value=header)
            if header in ['Sample ID(*)', 'Sample Code(*)', '样本编号(*)']:
                sample_id_col = i
            if header in ["DNB ID(*)"]:
                dnb_id_col = i
            if header in ["Index(*)", "Index号(*)"]:
                barcode_col = i

        self.samples_barcode = []
        self.dnb_id = "dnb_unit_test_%s" % (self.test_time.strftime("%m%d%H%M%S")) #str(uuid.uuid4())
        row = 2
        for i in range(1, number+1):
            serial_number = "sample_unit_test_%s_%s" %(self.test_time.strftime("%m%d%H%M%S"), str(i))
            for j in range(0, one_sample_barcode_number):
                barcode_number = BARCODE_LIST[row-2+BARCODE_START]
                self.samples.append(serial_number)
                self.samples_barcode.append({"subject_id": serial_number, "barcode_id": str(barcode_number)})
                for col, col_value in enumerate(table.row_values(1)):
                    sheet.cell(row=row, column=col+1, value=col_value)
                sheet.cell(row=row, column=sample_id_col+1, value=serial_number)
                sheet.cell(row=row, column=dnb_id_col+1, value=self.dnb_id)
                sheet.cell(row=row, column=barcode_col+1, value=str(barcode_number))
                row += 1
        wb.save(import_excel)
        r = self.zlims_client.callZlimsProImportDnbSample(import_excel)
        if r is False:
            raise Exception("Error: can't import dnb sample to zlims pro, error detail: %s" % r.get("msg"))

    def check_mom_service_status(self):
        self.mq_client = MqClient()
        client = self.mq_client.connect()
        if not client:
            raise Exception("Error: 连接不上ZLIMS activemq，请确认ZLIMS是否安装正常。")
        self.mq_client.disconnect()

    def start_listen_mq(self):
        logger.info("Starting listen MOM")
        self.mq_client = MqClient()
        self.mq_client.connect()
        if self.mq_client is None:
            raise Exception("Error: 连接不上ZLIMS activemq，请确认ZLIMS是否安装正常。")
        self.mq_listen_thread = threading.Thread(target=self.mq_client.subscribe_daemon, args=("/topic/*",))
        self.mq_listen_thread.start()

    def stop_listen_mq(self):
        logger.info("Stopping listen MOM")
        self.mq_client.stop_subscribe_daemon()
        self.mq_listen_thread.join()

    def check_dca_status(self):
        r = self.zlims_client.callZlimsApi("get_resource", {}, add_to_url="?part_number=DCA")
        if r:
            is_online = False
            for dca in r:
                instrument_status = dca["metadata"].get("mapped_instrument_status")
                self.need_check_values.update({"dca_software_version": dca["metadata"].get("software_version")})
                if instrument_status and instrument_status != "Offline":
                    is_online = True
                    break
            if not is_online and CHECK_INSTRUMENT_CURRENT_STATUS:
                raise Exception("Error: DCA服务器处于离线状态，请检查AMGR运行状态和网络连通性。")
        else:
            raise Exception("Error: DCA服务器没有连接上ZLIMS，请确认AMGR是否安装正常。")
        return

    def get_machine_id_for_T7(self):
        r = self.zlims_client.callZlimsApi("get_resource", {}, add_to_url="?part_number=%s" % T7)
        if r:
            self.machine_part_number = r[0]["part_number"]
            self.machine_sn = r[0]["serial_number"]
            self.machine_id = "%s,%s" % (self.machine_part_number, self.machine_sn)
            self.need_check_values.update({
                "basecall_version": r[0]["metadata"].get("basecall_version"),
                "control_software_version": r[0]["metadata"].get("control_software_version")
            })
            if r[0]["metadata"].get("mapped_instrument_status") == "Offline" and CHECK_INSTRUMENT_CURRENT_STATUS:
                raise Exception("Error: 仪器状态为离线状态，请检查仪器和ZLIMS服务器网络连接状态。")
        elif not CHECK_INSTRUMENT_CURRENT_STATUS:
            self.machine_part_number = T7
            self.machine_id = "%s,%s" % (self.machine_part_number, self.machine_sn)
            self.create_machine_instance()
        else:
            raise Exception("Error: 仪器没有连接上ZLIMS，请先将仪器连接上ZLIMS。")
        return r

    def get_machine_id(self):
        r = self.zlims_client.callZlimsApi("get_resource", {}, add_to_url="?serial_number=%s" % self.machine_sn)
        if r:
            self.machine_part_number = r[0]["part_number"]
            self.machine_id = "%s,%s" % (self.machine_part_number, self.machine_sn)
            self.need_check_values.update({
                "basecall_version": r[0]["metadata"].get("basecall_version"),
                "control_software_version": r[0]["metadata"].get("control_software_version")
            })
            if r[0]["metadata"].get("mapped_instrument_status") == "Offline" and CHECK_INSTRUMENT_CURRENT_STATUS:
                raise Exception("Error: 仪器状态为离线状态，请检查仪器和ZLIMS服务器网络连接状态。")
        else:
            raise Exception("Error: 仪器没有连接上ZLIMS，请先将仪器连接上ZLIMS。")
        return r

    def create_machine_instance(self):
        payload = self.zlims_client.create_machine_payload(self.machine_sn, self.machine_part_number)
        return self.zlims_client.callZlimsApi("create_resource", payload)

    def get_part_number_by_product(self):
        maps = {
            "QC": "QC",
            "BFI-TR": "BFI-Sample",
            "BFI-TB": "BFI-Sample"
        }
        part_number = maps.get(self.product, "%s-Sample" % self.product)
        return part_number

    def create_dnb(self):
        dnb_sn = str(uuid.uuid4())
        barcode_start = 1
        barcode_end = 10
        samples_barcode = []
        for barcode_number in range(barcode_start, barcode_end):
            sample_sn = str(uuid.uuid4())
            samples_barcode.append({"subject_id": sample_sn, "barcode_id": str(barcode_number)})
        ### create dnb instance
        payload = self.zlims_client.create_dnb_payload(dnb_sn, samples_barcode)
        r = self.zlims_client.callZlimsApi("create_resource", payload)
        self.assert_response(r, "Error: 创建DNB失败")

        self.samples_barcode = samples_barcode
        self.dnb_id = dnb_sn
        return dnb_sn

    def delete_dnb(self):
        dnb_sn = self.create_dnb()
        ### delete dnb instance
        payload = self.zlims_client.create_delete_dnb_payload(dnb_sn)
        r = self.zlims_client.callZlimsApi("delete_resource", payload)
        self.assert_response(r, "Error: 删除DNB失败")

    def get_analysis_warn_time(self):
        return get_analysis_time(self.product) + 2 * 60

    def get_sequencing_mode(self):
        maps = {}
        if self.sequencing_mode:
            return self.sequencing_mode
        sequencing_mode = maps.get(self.product, "PE")
        return sequencing_mode

    def get_lanes(self):
        maps = {
            "MGISEQ-200": ["L01"],
            "MGISEQ-2000": ["L01", "L02", "L03", "L04"]
        }
        lanes = maps.get(self.machine_part_number, ["L01"])
        if self.product in ONLY_SUPPORT_ONE_LANE:
            lanes = ["L01"]
        return lanes

    def _rsync_upload(self, fastq_path):
        if not CHECK_RSYNC_STATUS:
            return
        rsync_bat = os.path.join(sys.path[0], "mypackages", "cwRsync", "cwrsync.cmd")
        # self.rsync_config.get("rsync_bat")
        host = self.rsync_config.get("host")
        username = self.rsync_config.get("username")
        password = self.rsync_config.get("password")
        group = self.rsync_config.get("group")
        fastq_path = fastq_path.replace(':', '').replace('\\', '/')
        ### cwrsync.cmd 192.168.1.13 uploader halosuploader d/test uploader
        cmd = "\"%s\" %s %s %s %s %s" % (rsync_bat, host, username, password, fastq_path, group)
        fh = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, shell=True)
        try:
            content = fh.stdout.read()
            encode = chardet.detect(content).get("encoding")
            if encode:
                content = content.decode(encoding=encode)
            logger.info("==============upload logging==============")
            logger.info(content)
            logger.info("==========================================")
            search_obj = re.search(r'100%', content)
        except Exception as e:
            search_obj = None
            print("Error: %s" % str(e))
        if not search_obj and CHECK_RSYNC_STATUS:
            raise Exception("Error: 上传测试失败，请检查rsync服务器是否配置正常。\n 测试命令: %s" % cmd)

    def rsync_upload(self, fastq_path):
        self._rsync_upload(fastq_path)
        path = self.rsync_config.get("path")
        upload_path = "%s/%s/%s" % (path, self.machine_sn, self.flowcell_id)
        return upload_path.replace("//", "/")

    def test_rsync(self):
        # print("Test: rsync configure. please make sure has configure rsync at ZLIMS")
        test_path = os.path.join(sys.path[0], "test")
        test_file = os.path.join(test_path, "test.txt")
        ### test rsync install correctly
        if not os.path.exists(test_path):
            os.mkdir(test_path)
        fh = open(test_file, 'w')
        fh.write("unit test for whole pipeline %s" % time.time())
        fh.close()
        self._rsync_upload(test_path)

    def get_rsync_bat(self):
        rsync_bat = ""
        dir_names = ["Program Files (x86)", "Program Files"]
        all_software_names = ["ZebraV02Seq","ZebraV2Seq"]
        for dir_name in dir_names:
            for software_name in all_software_names:
                rsync_bat = "C:\\%s\\%s\\Service\\3rdParty\\cwRsync\\cwrsync.cmd"%(dir_name, software_name)
                if os.path.exists(rsync_bat):
                    return rsync_bat

        for dir_name in dir_names:
            for software_name in all_software_names:
                path = "C:\\%s\\%s" %(dir_name, software_name)
                for root, dirs, files in os.walk(path):
                    for file_name in files:
                        if file_name == "cwrsync.cmd":
                            return os.path.join(root, file_name)
        return rsync_bat

    def check_machine_config(self):
        ## 获取ZLIMS版本号
        r = self.zlims_client.callZlimsApi("get_version", {})
        self.assert_response(r)
        zlims_version = r.get("version_id")

        if self.machine_part_number == T7:
            self.need_check_values.update({
                "ZLIMS_version": zlims_version,
                "ZLIMS_host": self.zlims_client.zlims_host,
                "Instrument_sn": self.machine_sn
            })
            return
        ## 检查task是否配置正确
        instrument_config = self.parse_software_conf()
        config = instrument_config['ZLIMS']
        is_simulated = config.get('IsSimulated')
        is_online_mode = config.get("IsOnlineMode")
        # self.cal_upload_file = instrument_config['Basecall'].get("SyncUploadCalFilePath")
        if is_simulated == 'true' or is_simulated == 'True':
            raise Exception("Error: 控制软件配置了模拟模式，请修改好控制软件配置。")
        if is_online_mode == 'false' or is_online_mode == 'False':
            raise Exception("Error: 控制软件配置了离线模式，请修改好控制软件配置。")


        ## 检查rsync配置是否正确
        r = self.zlims_client.callZlimsApi("get_instrument_report", {}, add_to_url="?serial_number=%s" % self.machine_sn)
        self.assert_response(r)
        rsync_info = r[0]["rsync_info"]
        self.rsync_config = rsync_info
        self.need_check_values.update({
            "ZLIMS_version": zlims_version,
            "ZLIMS_host": self.zlims_client.zlims_host,
            "Rsync_host": rsync_info["host"],
            "Instrument_sn": self.machine_sn
        })

        self.test_rsync()

    def create_sample_by_resource_type(self, part_number, resource_type_instance):
        serial_number = str(uuid.uuid4())
        payload = {
            "part_number": part_number,
            "serial_number": serial_number,
            "metadata": {
                "product": self.product
            }
        }
        for field in resource_type_instance:
            validation_rule = field.get("validation_rule")
            required = field.get("required")
            field_name = field.get("field_name")
            type = field.get("type")
            value = field.get("value")
            if not value and required:
                if field_name in ["product"]:
                    continue
                if field_name == "host" and part_number == "PFI-Sample":
                    self.history_choice_dict[field_name] = {"key": "Human"}

                if isinstance(validation_rule, list) and len(validation_rule) > 0:
                    choice = self.history_choice_dict.get(field_name)
                    if choice is None:
                        choice = random.choice(validation_rule)
                        self.history_choice_dict[field_name] = choice
                    value = choice["key"]
                    if value == "RNA":
                        value = "DNA"
                else:
                    if type == "date":
                        value = datetime.datetime.utcnow().isoformat()
                    if type == "number":
                        value = random.randint(10)
                    else:
                        value = str(uuid.uuid4())
            payload["metadata"][field_name] = value
        r = self.zlims_client.callZlimsApi("create_resource", json.dumps(payload))
        self.assert_response(r)
        return serial_number

    def create_samples(self, number):
        part_number = self.get_part_number_by_product()
        r = self.zlims_client.callZlimsApi("get_drop_down_option", json.dumps({"part_number": part_number}))
        self.assert_response(r)
        for i in range(0, number):
            serial_number = self.create_sample_by_resource_type(part_number, r)
            self.samples.append(serial_number)

    def create_library_task(self):
        self.library_id = str(uuid.uuid4())
        barcode_number = 1
        samples = []
        for sample_sn in self.samples:
            samples.append({"subject_id": sample_sn, "barcode_id": str(barcode_number)})
            barcode_number += 1
        self.samples_barcode = samples
        start_payload = self.zlims_client.create_start_library_payload(self.machine_id, self.library_id, samples)
        complete_payload = self.zlims_client.create_complete_library_payload(samples)
        self.start_and_complete_task("library", start_payload, complete_payload)

    def create_pool_task(self):
        self.pool_id = str(uuid.uuid4())
        start_payload = self.zlims_client.create_start_pool_payload(self.machine_id, self.library_id, self.pool_id)
        complete_payload = self.zlims_client.create_complete_pool_payload(self.library_id, self.samples_barcode)
        self.start_and_complete_task("pool", start_payload, complete_payload)

    def create_make_dnb_task(self):
        self.dnb_id = str(uuid.uuid4())
        start_payload = self.zlims_client.create_start_make_dnb_payload(self.machine_id, self.pool_id, self.dnb_id)
        complete_payload = self.zlims_client.create_complete_make_dnb_payload()
        self.start_and_complete_task("make_dnb", start_payload, complete_payload)

    def create_load_dnb_task(self, dnb_id=None, flowcell_id=None):
        self.flowcell_id = "flowcellunittest%s" % (self.test_time.strftime("%m%d%H%M%S")) #str(uuid.uuid4())
        if flowcell_id:
            self.flowcell_id = flowcell_id
        lanes = self.get_lanes()
        new_lanes = []
        if dnb_id:
            r = self.zlims_client.callZlimsApi("get_barcode_info", {}, add_to_url="?dnb_id=%s" % dnb_id)
            self.assert_response(r, "Error: dnb id: %s 不存在，请先导入excel到ZLIMS Pro." % dnb_id)
            self.dnb_id = dnb_id
        for lane in lanes:
            new_lanes.append({
                "lane_id": lane,
                "dnb_id": self.dnb_id
            })
        start_payload = self.zlims_client.create_start_load_dnb_payload(new_lanes, self.flowcell_id, self.machine_id)
        update_payload = self.zlims_client.create_update_load_dnb_payload()
        complete_payload = self.zlims_client.create_complete_load_dnb_payload()
        task_inst_id = self.start_and_complete_task("load_dnb", start_payload, complete_payload, update_payload)
        miss_message_types = self.check_task_mq_messages(task_inst_id)
        if miss_message_types:
            raise Exception("Error: Load DNB task的 %s 消息获取失败，请联系ZLIMS团队检查消息接口" % miss_message_types)

    def check_task_mq_messages(self, task_inst_id, try_times=15):
        if not self.mq_client:
            return
        check_list = {
            "Task_Start": False,
            "Task_Update": False,
            "Task_Complete": False
        }
        miss_message_types = []
        for i in range(try_times):
            miss_message_types = []
            for message in self.mq_client.messages:
                try:
                    msg = json.loads(message)
                    msg_task_id = msg["message"]["task_inst_id"]
                    msg_type = msg.get("message_type")
                    if msg_task_id == task_inst_id and check_list.get(msg_type) is not None:
                        check_list[msg_type] = True
                except:
                    pass
            for key, value in check_list.items():
                if not value:
                    miss_message_types.append(key)
            if len(miss_message_types) == 0:
                break
            time.sleep(2)
        self.mq_client.clear_messages()
        if len(miss_message_types) > 0:
            return ",".join(miss_message_types)

    def check_sample_analysis_status(self, warning_time, error_exit_time):
        if not self.mq_client:
            return False

        samples_status = {}
        for sample in self.samples:
            samples_status[sample] = False
        start_check_time = datetime.datetime.utcnow()
        while True:
            for message in self.mq_client.messages:
                try:
                    msg = json.loads(message)
                    msg_dict = msg["message"]
                    msg_type = msg.get("message_type")
                    reports = msg_dict.get("inputs", {}).get("reports", [])
                    completion_status = msg_dict.get("completion_status")
                    if msg_type == "Task_Start":
                        notification_console("Info: Sample: %s, analysis task is running." % sn)
                    if msg_type == "Task_Complete":
                        notification_console("Info: Sample: %s, analysis task have completed, status is %s." %(sn, completion_status))
                    for report in reports:
                        analysis_result = report.get("analysis_result")
                        sn = report.get("serial_number")
                        report_html = report.get("report_html")

                        if analysis_result == 'success':
                            notification_console("Sample %s have completed. please check the report: %s is available." % (sn, report_html))
                            if sn in samples_status.keys():
                                samples_status[sn] = True
                except:
                    pass
            is_all_done = True
            for sample, status in samples_status.items():
                if not status:
                    is_all_done = False
                    break
            if is_all_done:
               break
            now = datetime.datetime.utcnow()
            delta = (now - start_check_time).total_seconds()
            if delta > error_exit_time:
                notification_console("Error: Analysis task analysis time is more than %ss. is maybe error, exit." %error_exit_time)
                return False
            self.mq_client.clear_messages()
            time.sleep(5)
        not_complete_samples = []
        for sample, status in samples_status.items():
            if not status:
                not_complete_samples.append(sample)
        return ",".join(not_complete_samples)

    def check_write_fastq_task(self, warning_time, error_exit_time):
        is_success = False
        start_check_time = datetime.datetime.utcnow()
        while True:
            for message in self.mq_client.messages:
                try:
                    msg = json.loads(message)
                    msg_dict = msg["message"]
                    msg_type = msg.get("message_type")
                    flowcell_id = msg_dict.get("inputs", {}).get("association_id")
                    task_id = msg_dict.get("task_id")
                    completion_status = msg_dict.get("completion_status")
                    if task_id == 'write_fastq' and self.flowcell_id == flowcell_id:
                        if msg_type == "Task_Start":
                            notification_console("Info: Flowcell ID: %s, write fastq task is running." % flowcell_id)
                        if msg_type == "Task_Complete":
                            notification_console("Info: Flowcell ID: %s, write fastq task have completed, status is %s." %(flowcell_id, completion_status))
                            if completion_status == 'success':
                                is_success = True
                            break
                except:
                    pass
            now = datetime.datetime.utcnow()
            delta = (now - start_check_time).total_seconds()
            if delta > error_exit_time:
                notification_console("Error: Flowcell ID: %s, write fastq task analysis time is more than %ss. is maybe error, exit." %(self.flowcell_id, error_exit_time))
                return False
            self.mq_client.clear_messages()
            time.sleep(5)
        return is_success

    def get_sample_barcode_from_zlims(self):
        if not hasattr(self, "samples_barcode"):
            self.samples_barcode = []
            r = self.zlims_client.callZlimsApi("get_barcode_info", {}, add_to_url="?dnb_id=%s" % self.dnb_id)
            self.assert_response(r)
            samples = r.get("samples")
            for sample in samples:
                self.samples_barcode.append({
                    "subject_id": sample.get("subject_id"),
                    "barcode_id": sample.get("barcode_id")
                })
                self.samples.append(sample.get("subject_id"))

    def get_barcode_info(self):
        self.get_sample_barcode_from_zlims()
        r = self.zlims_client.callZlimsApi("get_barcode_info", {}, add_to_url="?dnb_id=%s" % self.dnb_id)
        self.assert_response(r)
        samples = r.get("samples")
        for sample_barcode in self.samples_barcode:
            found_sample = None
            for sample in samples:
                if sample.get("subject_id") is None:
                    raise Exception("Error: get barcode info接口获取不到subject_id字段")
                if sample.get("barcode_id") is None:
                    raise Exception("Error: get barcode info接口获取不到barcode_id字段")
                if sample.get("atcg") is None:
                    raise Exception("Error: get barcode info接口获取不到atcg字段")
                if (sample["subject_id"] == sample_barcode["subject_id"] and
                            sample["barcode_id"] == sample_barcode["barcode_id"]):
                    found_sample = sample
                    break
            if not found_sample:
                raise Exception("Error: 在线下载样本barcode失败，获取不到sample信息。")
#            elif not found_sample.get('atcg'):
#                raise Exception("Error: 在线下载样本barcode失败，获取不到barcode序列")

        return r

    def create_sequencing_task(self):
        sequencing_mode = self.get_sequencing_mode()
        ## 测试获取barcode接口
        barcode_info = self.get_barcode_info()
        start_payload = self.zlims_client.create_start_sequencing_payload(sequencing_mode, self.flowcell_id, self.machine_id)
        r = self.zlims_client.callZlimsApi("start_task", start_payload)
        self.assert_response(r, "Error: 启动测序任务失败。")
        task_inst_id = r["task_inst_id"]

        lanes = self.get_lanes()
        for lane_id in lanes:
            for cycle_id in range(1, CYCLE_LENGTH+1):
                update_payload = self.zlims_client.create_update_sequencing_payload(lane_id, cycle_id)
                r = self.zlims_client.callZlimsApi("update_task", update_payload, format_to_url=task_inst_id)
                self.assert_response(r, "Error: 更新测序任务失败。")

        update_payload = self.zlims_client.create_update_write_fastq(lanes)
        r = self.zlims_client.callZlimsApi("update_task", update_payload, format_to_url=task_inst_id)
        self.assert_response(r, "Error: 更新测序任务失败。")

        sequencing_mode = self.get_sequencing_mode()
        if self.machine_part_number == T7:
            self.upload_cal_file()

        complete_payload = self.zlims_client.create_complete_sequencing_payload(lanes, barcode_info["samples"], sequencing_mode)
        r = self.zlims_client.callZlimsApi("complete_task", complete_payload, format_to_url=task_inst_id)
        self.assert_response(r, "Error: 结束测序任务失败。")
        miss_message_types = self.check_task_mq_messages(task_inst_id)
        if miss_message_types:
            raise Exception("Error: Sequencing task的 %s 消息获取失败，请联系ZLIMS团队检查消息接口" % miss_message_types)

    def test_write_fastq_task(self):
        analysis_time = get_analysis_time("write_fastq")
        is_success = self.check_write_fastq_task(analysis_time, analysis_time)
        if is_success is False:
            raise Exception("Error: Write task启动失败，请联系自动化分析团队，检查分析流程状态")

    def test_analysis_task_for_bio_pass(self):
        analysis_time = get_analysis_time(self.product)
        is_success = self.check_sample_analysis_status(analysis_time, analysis_time)
        if is_success is False:
            raise Exception("Error: %s分析任务启动失败，请联系自动化分析团队，检查分析流程状态" % self.product)

    def create_upload_task(self):
        lanes = self.get_lanes()
        rsync_host = self.rsync_config.get("host")
        upload_path = self.rsync_config.get("path")
        start_payload = self.zlims_client.create_start_upload_payload(lanes, self.flowcell_id, rsync_host, upload_path, self.machine_id)
        r = self.zlims_client.callZlimsApi("start_task", start_payload)
        self.assert_response(r, "Error: 创建上传任务失败。")
        task_inst_id = r["task_inst_id"]

        upload_path = self.upload_fastq()
        update_payload = self.zlims_client.create_update_upload_payload(lanes, upload_path)
        r = self.zlims_client.callZlimsApi("update_task", update_payload, format_to_url=task_inst_id)
        self.assert_response(r, "Error: 更新上传任务失败。")

        complete_payload = self.zlims_client.create_complete_upload_payload()
        r = self.zlims_client.callZlimsApi("complete_task", complete_payload, format_to_url=task_inst_id)
        self.assert_response(r, "Error: 结束上传任务失败。")

        analysis_task_id = r.get("analysis_task_inst_id")
        analysis_task_ids = r.get("analysis_task_inst_ids")
        if analysis_task_id:
            analysis_task_ids = [analysis_task_id]

        miss_message_types = self.check_task_mq_messages(task_inst_id)
        if miss_message_types:
            raise Exception("Error: Upload task的 %s 消息获取失败，请联系ZLIMS团队检查消息接口" % miss_message_types)

        self.analysis_task_ids = analysis_task_ids

    def _check_analysis_task_status(self, status, task_id, notification_time, message="Please check, something wrong."):
        start_time = datetime.datetime.utcnow()
        task_instance = None
        while True:
            task_instance = self.zlims_client.callZlimsApi("get_task_detail", {}, format_to_url=task_id)
            self.assert_response(task_instance, "Error: 获取不到分析任务。")
            task_status = task_instance["task_status"]
            if task_status in status:
                msg = "Info: Analysis task have %s." % task_status
                notification_console(msg)
                break
            else:
                now = datetime.datetime.utcnow()
                waited_seconds = (now - start_time).total_seconds()
                if waited_seconds > notification_time:
                    msg = "Warning: We have waited %ss, " % waited_seconds + message
                    notification_console(msg)
            time.sleep(30)
        return task_instance

    def test_analysis_task(self):
        if not self.analysis_task_ids or len(self.analysis_task_ids) < 1:
            raise Exception("Error: 不能自动创建分析任务。")

        for analysis_task_id in self.analysis_task_ids:
            ## 检查分析任务是否被启动
            notification_console("Starting check analysis task id: %s" % analysis_task_id)
            msg = "Analysis task is still not started, Please check DCA configuration and status."
            self._check_analysis_task_status(['started', 'completed'], analysis_task_id, ANALYSIS_TASK_START_WAIT_TIME, msg)

            ## 检查分析是否正常结束
            estimate_analysis_time = self.get_analysis_warn_time()
            msg = "Analysis task is still not completed, Please check DCA configuration and status."
            task_instance = self._check_analysis_task_status(['completed'], analysis_task_id, estimate_analysis_time, msg)
            if not task_instance or (task_instance and task_instance["completion_status"] != "success"):
                error_messages = task_instance.get("doc", {}).get("error_messages", '')
                search_obj = re.search('Child was return by signal', error_messages)
                msg = "Error: 分析失败。Error_Messages: %s" % error_messages
                if search_obj:
                    msg = "%s 分析软件异常退出，请联系生信团队排查问题。" % msg
                raise Exception(msg)
            reports = task_instance.get("doc", {}).get("context", {}).get('reports', [])
            analysis_pipeline_version = task_instance.get("doc", {}).get("context", {}).get('analysis_pipeline_version')
            if analysis_pipeline_version:
                self.need_check_values.update({"pipeline_version": analysis_pipeline_version})
            errors = self.check_reports(reports)
            if len(errors) > 0:
                raise Exception("Error: %s，请联系生信团队排查问题。" % ";".join(errors))

    def check_reports(self, reports):
        errors = []
        if not reports:
            errors.append("没有生成报告")
            return
        for report in reports:
            sn = report.get('serial_number')
            report_html = report.get('metadata', {}).get('report_html')
            if report_html is None:
                errors.append("%s 报告未生成" % sn)
        return errors

    def remove_temp_fastq(self, path):
        try:
            for root, dirs, files in os.walk(path):
                for file in files:
                    os.remove(os.path.join(root, file))
                    if file.endswith('fq.gz'):
                        try:
                            ## remove lane path
                            os.removedirs(root)
                            ## remove flowcell path
                            flowcell_path = os.path.dirname(root)
                            os.removedirs(flowcell_path)
                        except:
                            pass
        except Exception as e:
            print(str(e))

    def make_dir_for_basecall(self, test_dir):
        cmd = "sudo mkdir -p %s;sudo chown -R %s:%s %s" % (test_dir, basecall_uid, basecall_gid, test_dir)
        status = os.system(cmd)
        if status != 0:
            raise Exception("Error: can not make dir for write fastq")

    def cp_file_for_basecall(self, source, target):
        cmd = "sudo cp -fr %s %s;sudo chown -R %s:%s %s" % (source, target, basecall_uid, basecall_gid, target)
        status = os.system(cmd)
        if status != 0:
            raise Exception("Error: can not make file for write fastq")

    def get_flag_path(self, wdl_path):
        if not os.path.exists(wdl_path):
            wdl_path = "/storeData/ztron/apps/WDL_Install/wdl/write_fastq.workflow.wdl"
        with open(wdl_path, 'r') as fh:
            for line in fh.readlines():
                if line.strip().startswith("#"):
                    continue
                search_obj = re.search(r'String flag_dir = "([^"]+)"', line)
                if search_obj:
                    return search_obj.groups()[0]

                search_obj = re.search(r'flag_dir[\'"]: ?"([^"]+)"', line)
                if search_obj:
                    return search_obj.groups()[0]
        return None
   

    def upload_cal_file(self):
        test_flag_file = os.path.join(sys.path[0], TEST_CAL_FILE, "G10000004P_1_20190726_233352.json")
        test_cal_file = os.path.join(sys.path[0], TEST_CAL_FILE, "C035R035.cal")
        simulate_flag_file = "%s_1_20190726_233352.json" % self.flowcell_id
        flag_path = self.get_flag_path(WRITE_FASTQ_WDL_PATH)
        cal_upload_workspace_path = os.path.join(os.path.dirname(flag_path), "upload", "workspace")
        flag_file = os.path.join(flag_path, simulate_flag_file)
        if not os.path.exists(flag_path):
            self.make_dir_for_basecall(flag_path)
#            os.makedirs(flag_path)
        with open(test_flag_file, 'r') as fh:
            flag_json = json.loads(fh.read())
        flag_json["slide"] = self.flowcell_id
        flag_json["syncInfo"]["uploadPath"] = os.path.dirname(cal_upload_workspace_path)
        with open("test.flag", 'w') as fh:
            fh.write(json.dumps(flag_json))
        self.cp_file_for_basecall("test.flag", flag_file)
#        shutil.copyfile("test.flag", flag_file)
        
        cal_path = os.path.join(cal_upload_workspace_path, self.flowcell_id, "L01", "calFile")
        self.make_dir_for_basecall(cal_path)
#        os.makedirs(cal_path)
        metrics_path = os.path.join(cal_upload_workspace_path, self.flowcell_id, "L01", "metrics")
        self.make_dir_for_basecall(metrics_path)
        self.make_dir_for_basecall(os.path.join(cal_upload_workspace_path, self.flowcell_id))
#        os.makedirs(metrics_path)
        cal_file = os.path.join(cal_path, os.path.basename(test_cal_file))
        self.cp_file_for_basecall(test_cal_file, cal_file)
#        shutil.copyfile(test_cal_file, cal_file)

    def generate_flowcell_files(self, fastq_path_lane, flowcell, lane, barcode, sequencing_mode):
        sequence_stat = os.path.join(sys.path[0], TEST_FASTQ_PATH, "flowcell_files", "SequenceStat.txt")
        target_sequence_stat = os.path.join(fastq_path_lane, "SequenceStat.txt")
        basecall_stat = os.path.join(sys.path[0], TEST_FASTQ_PATH, "flowcell_files", "BarcodeStat.txt")
        target_basecall_stat = os.path.join(fastq_path_lane, "BarcodeStat.txt")
        summary_report = os.path.join(sys.path[0], TEST_FASTQ_PATH, "flowcell_files", "flowcell_id_L01.summaryReport.html")
        allfq_stat1 = os.path.join(sys.path[0], TEST_FASTQ_PATH, "flowcell_files", "read1.allfq.fqStat.txt")
        allfq_stat2 = os.path.join(sys.path[0], TEST_FASTQ_PATH, "flowcell_files", "read2.allfq.fqStat.txt")
        bfq_stat1 = os.path.join(sys.path[0], TEST_FASTQ_PATH, "flowcell_files", "read1.fq.fqStat.txt")
        bfq_stat2 = os.path.join(sys.path[0], TEST_FASTQ_PATH, "flowcell_files", "read2.fq.fqStat.txt")
        target_summary_report = os.path.join(fastq_path_lane, "%s_%s.summaryReport.html" % (flowcell, lane))
        origin_files = [sequence_stat, basecall_stat, summary_report, allfq_stat1, bfq_stat1]
        target_files = [target_sequence_stat, target_basecall_stat, target_summary_report]
        if sequencing_mode == "PE":
            target_allfq_stat1 = os.path.join(fastq_path_lane, "%s_%s_1.allfq.fqStat.txt" % (flowcell, lane))
            target_allfq_stat2 = os.path.join(fastq_path_lane, "%s_%s_2.allfq.fqStat.txt" % (flowcell, lane))
            target_bfq_stat1 = os.path.join(fastq_path_lane, "%s_%s_%s_1.fq.fqStat.txt" % (flowcell, lane, barcode))
            target_bfq_stat2 = os.path.join(fastq_path_lane, "%s_%s_%s_2.fq.fqStat.txt" % (flowcell, lane, barcode))
            origin_files.extend([allfq_stat2, bfq_stat2])
            target_files.extend([target_allfq_stat1, target_bfq_stat1, target_allfq_stat2, target_bfq_stat2])
        else:
            target_allfq_stat1 = os.path.join(fastq_path_lane, "%s_%s.allfq.fqStat.txt" % (flowcell, lane))
            target_bfq_stat1 = os.path.join(fastq_path_lane, "%s_%s_%s.fq.fqStat.txt" % (flowcell, lane, barcode))
            target_files.extend([target_allfq_stat1, target_bfq_stat1])
        for o, t in zip(origin_files, target_files):
            if os.path.exists(o):
                shutil.copyfile(o, t)

    def generate_fastq(self):
        sequencing_mode = self.get_sequencing_mode()
        read1 = os.path.join(sys.path[0], TEST_FASTQ_PATH, self.product, "read1.fq.gz")
        read2 = os.path.join(sys.path[0], TEST_FASTQ_PATH, self.product, "read2.fq.gz")
        ## 设置通用数据
        if not os.path.exists(read1) or ((not os.path.exists(read1) or not os.path.exists(read2)) and sequencing_mode == "PE"):
            read1 = os.path.join(sys.path[0], TEST_FASTQ_PATH, "common", "read1.fq.gz")
            read2 = os.path.join(sys.path[0], TEST_FASTQ_PATH, "common", "read2.fq.gz")

        self.remove_temp_fastq(os.path.join(sys.path[0], self.machine_sn))
        fastq_path = os.path.join(sys.path[0], self.machine_sn, self.flowcell_id)
        os.makedirs(fastq_path)
        lanes = self.get_lanes()
        index = 0
        for lane in lanes:
            fastq_path_lane = os.path.join(fastq_path, lane)
            os.makedirs(fastq_path_lane)
            for sample_barcode in self.samples_barcode:
                barcode_id = sample_barcode["barcode_id"]
                ### try get different fastq file
                new_read1_file = "%s__%s" %(read1.split('__')[0], index)
                if index > 0 and os.path.exists(new_read1_file):
                    read1 = new_read1_file
                new_read2_file = "%s__%s" %(read2.split('__')[0], index)
                if index > 0 and os.path.exists(new_read2_file):
                    read2 = new_read2_file
                if sequencing_mode == "PE":
                    if not os.path.exists(read1) or not os.path.exists(read2):
                        raise Exception("Error: 找不到测试fastq文件。")
                    read1_file = os.path.join(fastq_path_lane,
                                                     "%s_%s_%s_1.fq.gz" % (self.flowcell_id, lane, barcode_id))
                    shutil.copyfile(read1, read1_file)
                    read2_file = os.path.join(fastq_path_lane,
                                                     "%s_%s_%s_2.fq.gz" % (self.flowcell_id, lane, barcode_id))
                    shutil.copyfile(read2, read2_file)
                else:
                    if not os.path.exists(read1):
                        raise Exception("Error: 找不到测试fastq文件。")
                    read1_file = os.path.join(fastq_path_lane,
                                                     "%s_%s_%s.fq.gz" % (self.flowcell_id, lane, barcode_id))
                    shutil.copyfile(read1, read1_file)
                index += 1
                try:
                    self.generate_flowcell_files(fastq_path_lane, self.flowcell_id, lane, barcode_id, sequencing_mode)
                except:
                    pass
        return os.path.join(sys.path[0], self.machine_sn)

    def upload_fastq(self):
        fastq_path = self.generate_fastq()
        upload_path = self.rsync_upload(fastq_path)
        return upload_path

    def start_and_complete_task(self, task_name, start_payload, complete_payload, update_payload=None):
        r = self.zlims_client.callZlimsApi("start_task", start_payload)
        self.assert_response(r, "Error: 创建 %s 任务失败" % task_name)
        task_inst_id = r["task_inst_id"]
        if update_payload:
            r = self.zlims_client.callZlimsApi("update_task", update_payload, format_to_url=task_inst_id)
            self.assert_response(r, "Error: 更新 %s 任务失败" % task_name)
        r = self.zlims_client.callZlimsApi("complete_task", complete_payload, format_to_url=task_inst_id)
        self.assert_response(r, "Error: 结束 %s 任务失败" % task_name)
        return task_inst_id

    def assert_response(self, response, message="Call ZLIMS API error!"):
        if response is False:
            raise ZLIMSServiceError(message)


def test_create_samples():
    products = "QC|WGS|WES|BFI|PFI|HMBI|BFI-TR|BFI-TB"
    for product in products.split("|"):
        pipeline_test = WholePipelineTest(product)
        pipeline_test.create_samples(1)
        print(pipeline_test.samples)


if __name__ == '__main__':
    pipeline_test_tools = WholePipelineTest(TEST_PRODUCTS[0])
    pipeline_test_tools.check_dca_status()
    pipeline_test_tools.get_machine_id()
    pipeline_test_tools.check_machine_config()
    pipeline_test_tools.create_samples(8)
    pipeline_test_tools.create_library_task()
    pipeline_test_tools.create_pool_task()
    pipeline_test_tools.create_make_dnb_task()
    pipeline_test_tools.create_load_dnb_task()
    pipeline_test_tools.create_sequencing_task()
    pipeline_test_tools.create_upload_task()
    pipeline_test_tools.test_analysis_task()
